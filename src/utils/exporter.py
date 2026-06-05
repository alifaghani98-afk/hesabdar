"""
exporter.py - Excel export using openpyxl directly (no pandas/numpy dependency).
"""
import os
import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

from ..db.transactions_dao import get_transactions
from .jalali_utils import format_jalali_display

EXPORT_DIR = os.path.join(os.path.expanduser("~"), "Documents", "HesabdarExports")


def _ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)


def _timestamp_str():
    return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


def _write_excel(transactions, filepath, sheet_name="transactions"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.sheet_view.rightToLeft = True

    headers = ['row', 'date', 'description', 'debit', 'credit', 'balance', 'project', 'counterparty']
    header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    running_balance = 0.0
    total_debit = 0.0
    total_credit = 0.0

    for idx, t in enumerate(transactions, start=1):
        running_balance += t['credit'] - t['debit']
        total_debit += t['debit']
        total_credit += t['credit']

        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=format_jalali_display(t['jalali_date']))
        ws.cell(row=row, column=3, value=t['description'])
        ws.cell(row=row, column=4, value=t['debit'] if t['debit'] > 0 else '')
        ws.cell(row=row, column=5, value=t['credit'] if t['credit'] > 0 else '')
        ws.cell(row=row, column=6, value=running_balance)
        ws.cell(row=row, column=7, value=t.get('project_name', ''))
        ws.cell(row=row, column=8, value=t.get('counterparty_name', ''))

        if t['debit'] > 0:
            ws.cell(row=row, column=4).font = Font(color="CC0000")
        if t['credit'] > 0:
            ws.cell(row=row, column=5).font = Font(color="007700")

    last_row = len(transactions) + 3
    ws.cell(row=last_row,     column=1, value="Total Debit:")
    ws.cell(row=last_row,     column=2, value=total_debit)
    ws.cell(row=last_row + 1, column=1, value="Total Credit:")
    ws.cell(row=last_row + 1, column=2, value=total_credit)
    ws.cell(row=last_row + 2, column=1, value="Balance:")
    ws.cell(row=last_row + 2, column=2, value=total_credit - total_debit)

    for r in [last_row, last_row+1, last_row+2]:
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).font = Font(bold=True)

    widths = [6, 14, 35, 14, 14, 16, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(filepath)


def export_by_project(project_id, project_name):
    if not _OPENPYXL_OK:
        raise RuntimeError("openpyxl not found.")
    _ensure_export_dir()
    txs = get_transactions(project_id=project_id)
    safe = project_name.replace('/', '-')
    path = os.path.join(EXPORT_DIR, f"project_{safe}_{_timestamp_str()}.xlsx")
    _write_excel(txs, path, sheet_name=project_name[:31])
    return path


def export_by_counterparty(counterparty_id, counterparty_name):
    if not _OPENPYXL_OK:
        raise RuntimeError("openpyxl not found.")
    _ensure_export_dir()
    txs = get_transactions(counterparty_id=counterparty_id)
    safe = counterparty_name.replace('/', '-')
    path = os.path.join(EXPORT_DIR, f"counterparty_{safe}_{_timestamp_str()}.xlsx")
    _write_excel(txs, path, sheet_name=counterparty_name[:31])
    return path


def export_full_system():
    if not _OPENPYXL_OK:
        raise RuntimeError("openpyxl not found.")
    _ensure_export_dir()
    txs = get_transactions()
    path = os.path.join(EXPORT_DIR, f"full_export_{_timestamp_str()}.xlsx")
    _write_excel(txs, path, sheet_name="all_transactions")
    return path
